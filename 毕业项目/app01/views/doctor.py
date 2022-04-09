from django.shortcuts import render, redirect
from app01 import models
from app01.utils.pagination import Pagination
from app01.utils.form import DoctorModelForm
from django.http import JsonResponse
from openpyxl import load_workbook


def list(request):
    """医生列表"""
    data_dict = {}
    # 没有值则填入默认值""
    search_data = request.GET.get('Search_doctor', "")
    if search_data:
        data_dict["name__contains"] = search_data
    # select * from 表 order by level asc

    queryset = models.Doctor.objects.filter(**data_dict).order_by("level").order_by("depart")

    page_object = Pagination(request, queryset, page_size=15)

    context = {"search_data": search_data, "queryset": page_object.page_queryset, "page_string": page_object.html()}
    return render(request, "doctor_list.html", context)


def add(request):
    """添加医生"""
    if request.method == "GET":
        form = DoctorModelForm()
        return render(request, "doctor_add.html", {"form": form})
    # 获取数据，并进行校验
    form = DoctorModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        return redirect('/doctor/list/')
    return render(request, "doctor_add.html", {"form": form})


def excel(request):
    """批量上传（EXCEl文件),仅适用与xlxs文件"""
    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    for row in sheet.iter_rows(min_row=2):
        name = row[0].value
        exists = models.Doctor.objects.filter(name=name).exists()
        gender = row[1].value
        depart = row[2].value
        level = row[3].value
        salary = row[4].value
        day = row[5].value
        # 判断姓名是否存在，不存在则新增，存在则更新
        if not exists:
            models.Doctor.objects.create(name=name, gender=gender, depart_id=depart, level=level, salary=salary,
                                         day=day)
        else:
            models.Doctor.objects.filter(name=name).update(name=name, gender=gender, depart_id=depart, level=level,
                                                           salary=salary,day=day)
    return redirect("/doctor/list/")


def edit(request, nid):
    """编辑医生"""
    row_object = models.Doctor.objects.filter(id=nid).first()

    if request.method == "GET":
        form = DoctorModelForm(instance=row_object)
        return render(request, 'doctor_edit.html', {"form": form})

    form = DoctorModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/doctor/list/')
    return render(request, 'doctor_edit.html', {"form": form})


def delete(request):
    """删除医生"""
    uid = request.GET.get("uid")
    models.Doctor.objects.filter(id=uid).delete()

    return JsonResponse({"status": True})
