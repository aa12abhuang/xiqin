from django.shortcuts import render, redirect
from app01 import models
from app01.utils.pagination import Pagination
from app01.utils.form import MedicineModelForm
from django.http import JsonResponse
from openpyxl import load_workbook


def list(request):
    """药物列表"""
    data_dict = {}
    # 没有值则填入默认值""
    search_data = request.GET.get('Search_medicine', "")
    if search_data:
        data_dict["name__contains"] = search_data
    # select * from 表 order by level asc

    queryset = models.Medicine.objects.filter(**data_dict)

    page_object = Pagination(request, queryset, page_size=15)

    context = {"search_data": search_data, "queryset": page_object.page_queryset, "page_string": page_object.html()}
    return render(request, "medicine_list.html", context)


def add(request):
    """添加药物"""
    if request.method == "GET":
        form = MedicineModelForm()
        return render(request, "medicine_add.html", {"form": form})
    # 获取数据，并进行校验
    form = MedicineModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        return redirect('/medicine/list/')
    return render(request, "medicine_add.html", {"form": form})


def excel(request):
    """批量上传（EXCEl文件),仅适用与xlxs文件"""
    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    for row in sheet.iter_rows(min_row=2):
        name = row[0].value
        exists = models.Medicine.objects.filter(name=name).exists()
        scale = row[1].value
        unit = row[2].value
        price = row[3].value
        category = row[4].value
        # 判断姓名是否存在，不存在则新增，存在则更新
        if not exists:
            models.Medicine.objects.create(name=name, scale=scale, unit=unit, price=price, category=category)
        else:
            models.Medicine.objects.filter(name=name).update(name=name, scale=scale, unit=unit, price=price,
                                                             category=category)
    return redirect("/medicine/list/")


def edit(request, nid):
    """编辑药物"""
    row_object = models.Medicine.objects.filter(id=nid).first()

    if request.method == "GET":
        form = MedicineModelForm(instance=row_object)
        return render(request, 'medicine_edit.html', {"form": form})

    form = MedicineModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/medicine/list/')
    return render(request, 'medicine_edit.html', {"form": form})


def delete(request):
    """删除医生"""
    uid = request.GET.get("uid")
    models.Medicine.objects.filter(id=uid).delete()

    return JsonResponse({"status": True})